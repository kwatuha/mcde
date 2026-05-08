#!/usr/bin/env python3
"""
Script to transform budget source file into the correct import format.
Reads from 2025_2026_budgets_source.xlsx and outputs to budget_mapping_template_import_now.xlsx
"""

import pandas as pd
import subprocess
import re
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Optional

# Database connection details - try gov_db first, then kisumu_db, then db
DB_CONTAINER = None  # Will be determined dynamically
DB_NAME = "gov_imbesdb"
DB_USER = "root"

def get_db_container():
    """Determine the correct database container name."""
    # Try gov_db first (for government_projects branch)
    result = subprocess.run(
        ["docker", "ps", "--filter", "name=gov_db", "--format", "{{.Names}}"],
        capture_output=True,
        text=True
    )
    if result.stdout.strip():
        return "gov_db"
    
    # Try kisumu_db
    result = subprocess.run(
        ["docker", "ps", "--filter", "name=kisumu_db", "--format", "{{.Names}}"],
        capture_output=True,
        text=True
    )
    if result.stdout.strip():
        return "kisumu_db"
    
    # Try db
    result = subprocess.run(
        ["docker", "ps", "--filter", "name=^db$", "--format", "{{.Names}}"],
        capture_output=True,
        text=True
    )
    if result.stdout.strip():
        return "db"
    
    # Try any mysql container
    result = subprocess.run(
        ["docker", "ps", "--filter", "ancestor=mysql", "--format", "{{.Names}}"],
        capture_output=True,
        text=True
    )
    if result.stdout.strip():
        return result.stdout.strip().split('\n')[0]
    
    raise Exception("No MySQL container found")

def get_db_password():
    """Get MySQL root password from docker container."""
    global DB_CONTAINER
    if DB_CONTAINER is None:
        DB_CONTAINER = get_db_container()
    
    try:
        result = subprocess.run(
            ["docker", "exec", DB_CONTAINER, "printenv", "MYSQL_ROOT_PASSWORD"],
            capture_output=True,
            text=True,
            check=True
        )
        return result.stdout.strip() or "root"
    except:
        return "root"

def query_database(query: str) -> List[Dict]:
    """Execute MySQL query and return results as list of dicts."""
    global DB_CONTAINER
    if DB_CONTAINER is None:
        DB_CONTAINER = get_db_container()
        print(f"Using database container: {DB_CONTAINER}")
    
    password = get_db_password()
    cmd = [
        "docker", "exec", "-i", DB_CONTAINER,
        "mysql", "-u", DB_USER, f"-p{password}", DB_NAME,
        "-e", query
    ]
    
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            check=True
        )
        # Parse tab-separated output
        lines = result.stdout.strip().split('\n')
        if len(lines) < 2:
            return []
        
        headers = lines[0].split('\t')
        data = []
        for line in lines[1:]:
            if line.strip():
                values = line.split('\t')
                if len(values) == len(headers):
                    data.append(dict(zip(headers, values)))
        return data
    except subprocess.CalledProcessError as e:
        print(f"Database query error: {e.stderr}")
        return []

def load_database_mappings() -> Tuple[Dict, Dict, Dict]:
    """Load departments, wards, and subcounties from database."""
    print("Loading database mappings...")
    
    # Load departments (only voided = 0)
    dept_query = "SELECT departmentId, name FROM departments WHERE voided = 0;"
    dept_data = query_database(dept_query)
    departments = {}
    for row in dept_data:
        dept_name = row.get('name', '').strip()
        dept_id = row.get('departmentId', '').strip()
        if dept_name:
            # Store normalized version for matching
            normalized = normalize_text(dept_name)
            departments[normalized] = {'id': dept_id, 'name': dept_name}
            # Also store without "Department:" prefix if present
            if normalized.startswith('department:'):
                departments[normalized.replace('department:', '').strip()] = {'id': dept_id, 'name': dept_name}
    
    # Load wards and subcounties (only voided = 0)
    ward_query = """
    SELECT w.wardId, w.name as wardName, sc.subcountyId, sc.name as subcountyName 
    FROM wards w 
    LEFT JOIN subcounties sc ON w.subcountyId = sc.subcountyId 
    WHERE w.voided = 0;
    """
    ward_data = query_database(ward_query)
    wards = {}
    subcounties = {}
    
    for row in ward_data:
        ward_name = row.get('wardName', '').strip()
        subcounty_name = row.get('subcountyName', '').strip()
        ward_id = row.get('wardId', '').strip()
        subcounty_id = row.get('subcountyId', '').strip()
        
        if ward_name:
            # Store normalized ward name
            normalized = normalize_text(ward_name)
            wards[normalized] = {
                'id': ward_id,
                'name': ward_name,
                'subcountyId': subcounty_id,
                'subcountyName': subcounty_name
            }
            # Also store variations (with/without quotes, different quote types)
            # Handle "Nyalenda A" vs "Nyalenda 'A'" vs "Nyalenda \"A\""
            normalized_no_quotes = re.sub(r'["\']', '', normalized).strip()
            if normalized_no_quotes != normalized:
                wards[normalized_no_quotes] = {
                    'id': ward_id,
                    'name': ward_name,
                    'subcountyId': subcounty_id,
                    'subcountyName': subcounty_name
                }
        
        if subcounty_name:
            normalized = normalize_text(subcounty_name)
            subcounties[normalized] = {
                'id': subcounty_id,
                'name': subcounty_name
            }
    
    print(f"Loaded {len(departments)} departments, {len(wards)} wards, {len(subcounties)} subcounties")
    return departments, wards, subcounties

def normalize_text(text: str) -> str:
    """Normalize text for matching - remove extra spaces, lowercase, etc."""
    if not text or pd.isna(text):
        return ""
    # Convert to string, strip, lowercase
    normalized = str(text).strip().lower()
    # Remove extra spaces
    normalized = re.sub(r'\s+', ' ', normalized)
    return normalized

def find_matching_department(dept_name: str, departments: Dict) -> Optional[Dict]:
    """Find matching department in database."""
    if not dept_name:
        return None
    
    normalized = normalize_text(dept_name)
    
    # Try exact match
    if normalized in departments:
        return departments[normalized]
    
    # Try removing "Department:" prefix
    cleaned = re.sub(r'^(department:|dept\.?|dept\s+)', '', normalized, flags=re.IGNORECASE).strip()
    if cleaned in departments:
        return departments[cleaned]
    
    # Try partial matches (contains)
    for db_dept, dept_info in departments.items():
        if normalized in db_dept or db_dept in normalized:
            return dept_info
    
    # Try fuzzy matching - check if key words match
    normalized_words = set(normalized.split())
    for db_dept, dept_info in departments.items():
        db_words = set(db_dept.split())
        # If most words match (at least 50% overlap)
        if len(normalized_words) > 0 and len(db_words) > 0:
            overlap = len(normalized_words & db_words)
            if overlap >= min(len(normalized_words), len(db_words)) * 0.5:
                return dept_info
    
    return None

def find_matching_ward(ward_name: str, wards: Dict) -> Optional[Dict]:
    """Find matching ward in database."""
    if not ward_name:
        return None
    
    # Handle special cases first - "All Wards" or "All Ward" -> CountyWide
    normalized = normalize_text(ward_name)
    normalized_clean = re.sub(r'[-\s]+', ' ', normalized).strip()
    
    if normalized_clean in ['all wards', 'all ward']:
        return {'name': 'CountyWide', 'subcountyName': 'CountyWide', 'isCountyWide': True}
    
    # Check if it contains "all" and "ward" (in any order)
    words = set(normalized_clean.split())
    if 'all' in words and 'ward' in words:
        return {'name': 'CountyWide', 'subcountyName': 'CountyWide', 'isCountyWide': True}
    
    # Handle "countywide" variations
    if 'countywide' in normalized or 'county wide' in normalized_clean:
        return {'name': 'CountyWide', 'subcountyName': 'CountyWide', 'isCountyWide': True}
    
    # Handle compound ward names (e.g., "Kisumu East and Kisumu Central")
    # For compound names, try to match the first part
    if ' and ' in normalized_clean:
        parts = normalized_clean.split(' and ')
        # Try to match the first part
        first_part = parts[0].strip()
        match = find_matching_ward(first_part, wards)
        if match:
            return match
        # If first part doesn't match, return None (can't map compound wards)
        return None
    
    # Remove quotes for matching (e.g., "Nyalenda A" -> Nyalenda A)
    normalized_no_quotes = re.sub(r'["\']', '', normalized).strip()
    
    # Normalize spaces and special characters for matching
    normalized_clean = re.sub(r'["\']', '', normalized).replace('/', ' ').replace('-', ' ').replace('  ', ' ').strip()
    
    # Try exact match
    if normalized in wards:
        return wards[normalized]
    
    # Try match without quotes
    if normalized_no_quotes in wards:
        return wards[normalized_no_quotes]
    
    # Try matching with normalized spaces
    if normalized_clean in wards:
        return wards[normalized_clean]
    
    # Try matching with variations (handle "Nyalenda \"A\"" vs "Nyalenda A" vs "Nyalenda 'A'")
    for db_ward, ward_info in wards.items():
        db_ward_clean = re.sub(r'["\']', '', db_ward).replace('/', ' ').replace('-', ' ').replace('  ', ' ').strip()
        if normalized_clean == db_ward_clean:
            return ward_info
    
    # Try partial matches (handle cases like "Kisumu East" matching "EAST KISUMU" or "KISUMU EAST")
    normalized_words = set(normalized_clean.split())
    for db_ward, ward_info in wards.items():
        db_ward_clean = re.sub(r'["\']', '', db_ward).replace('/', ' ').replace('-', ' ').replace('  ', ' ').strip()
        db_words = set(db_ward_clean.split())
        # If all words match (order independent)
        if normalized_words == db_words and len(normalized_words) > 0:
            return ward_info
    
    # Try word-by-word matching (for compound names or partial matches)
    # Check if all words from normalized are in db_ward (or vice versa)
    for db_ward, ward_info in wards.items():
        db_ward_clean = re.sub(r'["\']', '', db_ward).replace('/', ' ').replace('-', ' ').replace('  ', ' ').strip()
        db_words = set(db_ward_clean.split())
        # If normalized words are a subset of db_words or vice versa (and at least 2 words match)
        if len(normalized_words) > 0 and len(db_words) > 0:
            overlap = normalized_words & db_words
            if len(overlap) >= min(2, len(normalized_words), len(db_words)):
                return ward_info
    
    # Final attempt: Try matching individual significant words (ignore common words like "and", "the", etc.)
    # For "Kisumu East", try to find wards containing both "kisumu" and "east"
    significant_words = {w for w in normalized_words if w not in ['and', 'the', 'of', 'in', 'on', 'at', 'to', 'for']}
    if len(significant_words) >= 2:
        for db_ward, ward_info in wards.items():
            db_ward_clean = re.sub(r'["\']', '', db_ward).replace('/', ' ').replace('-', ' ').replace('  ', ' ').strip()
            db_words = set(db_ward_clean.split())
            # Check if all significant words are present in db_ward
            if significant_words.issubset(db_words):
                return ward_info
    
    return None

def extract_data_from_source(df: pd.DataFrame) -> List[Dict]:
    """Extract project data from source DataFrame."""
    data = []
    current_department = None
    
    # Find header row (contains "S/No" or "Project")
    header_row_idx = None
    for idx, row in df.iterrows():
        row_str = ' '.join([str(cell) for cell in row.values if pd.notna(cell)]).lower()
        if 's/no' in row_str or 'project' in row_str:
            header_row_idx = idx
            break
    
    # Extract department from first row if present
    if len(df) > 0:
        first_row = df.iloc[0]
        for col in df.columns:
            val = str(first_row[col]) if pd.notna(first_row[col]) else ""
            if 'department:' in val.lower():
                match = re.search(r'department:\s*(.+)', val, re.IGNORECASE)
                if match:
                    current_department = match.group(1).strip()
                    break
    
    if header_row_idx is None:
        print("  Warning: Could not find header row, using row 1 as header")
        header_row_idx = 1
    
    # Get headers
    headers = df.iloc[header_row_idx]
    
    # Find column indices
    sno_col = None
    project_col = None
    ward_col = None
    amount_col = None
    
    for idx, header in enumerate(headers):
        header_str = str(header).lower() if pd.notna(header) else ""
        if ('s/no' in header_str or 'sno' in header_str or 'serial' in header_str) and sno_col is None:
            sno_col = idx
        elif 'project' in header_str and project_col is None:
            project_col = idx
        elif 'ward' in header_str and ward_col is None:
            ward_col = idx
        elif 'amount' in header_str and amount_col is None:
            amount_col = idx
    
    # If columns not found, try common positions
    if sno_col is None:
        sno_col = 0  # S/N is usually first column
    if project_col is None:
        project_col = 1 if len(df.columns) > 1 else 0
    if ward_col is None:
        ward_col = 2 if len(df.columns) > 2 else 1
    if amount_col is None:
        amount_col = 3 if len(df.columns) > 3 else 2
    
    # Extract data rows
    for idx in range(header_row_idx + 1, len(df)):
        row = df.iloc[idx]
        
        sno = row.iloc[sno_col] if pd.notna(row.iloc[sno_col]) and sno_col < len(row) else ""
        project = str(row.iloc[project_col]).strip() if pd.notna(row.iloc[project_col]) and project_col < len(row) else ""
        ward = str(row.iloc[ward_col]).strip() if pd.notna(row.iloc[ward_col]) and ward_col < len(row) else ""
        amount = row.iloc[amount_col] if pd.notna(row.iloc[amount_col]) and amount_col < len(row) else None
        
        # Skip empty rows
        if not project or project.lower() in ['s/no', 'project', 'ward', 'amount', 'nan', '']:
            continue
        
        # Skip if amount is not numeric
        try:
            amount_float = float(amount) if amount is not None else 0
        except (ValueError, TypeError):
            continue
        
        data.append({
            'sno': sno,
            'project': project,
            'ward': ward,
            'amount': amount_float,
            'department': current_department
        })
    
    return data

def process_budget_file(source_file: str, output_file: str):
    """Process source budget file and create output in template format."""
    print(f"Reading source file: {source_file}")
    
    # Read all sheets from the Excel file
    xls = pd.ExcelFile(source_file)
    print(f"Found {len(xls.sheet_names)} sheet(s): {xls.sheet_names}")
    
    # Load database mappings
    departments, wards, subcounties = load_database_mappings()
    
    # Extract data from all sheets
    print("\nExtracting data from all sheets...")
    all_data = []
    current_department = None
    
    for sheet_name in xls.sheet_names:
        print(f"\nProcessing sheet: {sheet_name}")
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        
        # Extract data from this sheet
        sheet_data = extract_data_from_source(df)
        
        # Update current_department from the sheet if found
        if len(df) > 0:
            first_row = df.iloc[0]
            for col in df.columns:
                val = str(first_row[col]) if pd.notna(first_row[col]) else ""
                if 'department:' in val.lower():
                    match = re.search(r'department:\s*(.+)', val, re.IGNORECASE)
                    if match:
                        current_department = match.group(1).strip()
                        print(f"  Found department: {current_department}")
                        break
        
        # Update department for all items in this sheet if we found one
        if current_department:
            for item in sheet_data:
                if not item.get('department'):
                    item['department'] = current_department
        
        print(f"  Extracted {len(sheet_data)} items from sheet '{sheet_name}'")
        all_data.extend(sheet_data)
    
    print(f"\nTotal items extracted from all sheets: {len(all_data)}")
    
    # Create output DataFrame
    output_data = []
    unmatched_departments = set()
    unmatched_wards = set()
    
    for item in all_data:
        # Find matching department
        dept_match = find_matching_department(item['department'], departments)
        db_department = dept_match['name'] if dept_match else None
        if not db_department:
            unmatched_departments.add(item['department'])
            db_department = 'unknown'  # Set to 'unknown' if not matched
        
        # Find matching ward and subcounty
        ward_match = find_matching_ward(item['ward'], wards)
        if ward_match:
            if ward_match.get('isCountyWide'):
                db_ward = 'CountyWide'
                db_subcounty = 'CountyWide'
            else:
                db_ward = ward_match.get('name', 'unknown')
                db_subcounty = ward_match.get('subcountyName', 'unknown')
        else:
            db_ward = "unknown"
            db_subcounty = "unknown"
            unmatched_wards.add(item['ward'])
        
        output_data.append({
            'S/N': item.get('sno', ''),  # Serial number from source file
            'Budget': 'Approved Budget FY 2025/2026',
            'Project Name': item['project'],
            'Amount': item['amount'],
            'ward': db_ward,
            'subcounty': db_subcounty,
            'fin_year': '2025/2026',
            'db_department': db_department,  # Matched department name from departments
            'original_ward': item['ward'],  # Original ward from source file
            'original_department': item['department']  # Original department from source file
        })
    
    # Create DataFrame
    output_df = pd.DataFrame(output_data)
    
    # Write to output file
    print(f"\nWriting to output file: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='Sheet1', index=False)
        
        # Get the worksheet to adjust column widths and enable wrapping
        worksheet = writer.sheets['Sheet1']
        
        # Set column widths
        column_widths = {
            'S/N': 8,
            'Budget': 30,
            'Project Name': 60,
            'Amount': 15,
            'ward': 25,
            'subcounty': 30,
            'fin_year': 15,
            'db_department': 50,  # Matched department from database
            'original_ward': 30,
            'original_department': 50
        }
        
        column_names = list(output_df.columns)
        for idx, col_name in enumerate(column_names, start=1):
            col_letter = get_column_letter(idx)
            
            # Calculate max width based on content
            max_length = column_widths.get(col_name, 20)
            for value in output_df[col_name]:
                if pd.notna(value):
                    cell_length = len(str(value))
                    if cell_length > max_length:
                        max_length = cell_length
            
            # Set column width (add some padding, max 100 characters)
            adjusted_width = min(max_length + 2, 100)
            worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Enable text wrapping for all cells in this column
            for row in range(2, len(output_df) + 2):  # Start from row 2 (skip header)
                cell = worksheet[f'{col_letter}{row}']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Also wrap header row
        for idx, col_name in enumerate(column_names, start=1):
            col_letter = get_column_letter(idx)
            header_cell = worksheet[f'{col_letter}1']
            header_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
            header_cell.font = Font(bold=True)
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    print(f"Successfully created output file with {len(output_df)} rows")
    print(f"\nSummary:")
    print(f"  Total rows: {len(output_df)}")
    print(f"  Wards matched: {len([d for d in output_data if d['ward'] not in ['unknown', 'CountyWide']])}")
    print(f"  Subcounties matched: {len([d for d in output_data if d['subcounty'] not in ['unknown', 'CountyWide']])}")
    print(f"  CountyWide entries: {len([d for d in output_data if d['ward'] == 'CountyWide'])}")
    print(f"  Unknown wards: {len([d for d in output_data if d['ward'] == 'unknown'])}")
    print(f"  Unknown subcounties: {len([d for d in output_data if d['subcounty'] == 'unknown'])}")
    
    if unmatched_departments:
        print(f"\n  Unmatched departments ({len(unmatched_departments)}):")
        for dept in sorted(unmatched_departments):
            print(f"    - {dept}")
    
    if unmatched_wards:
        print(f"\n  Unmatched wards ({len(unmatched_wards)}):")
        for ward in sorted(unmatched_wards):
            print(f"    - {ward}")

if __name__ == "__main__":
    source_file = "/home/dev/dev/imes_working/v5/budgets/2025_2026_budgets_source.xlsx"
    output_file = "/home/dev/dev/imes_working/v5/budgets/budget_mapping_template_import_now.xlsx"
    
    process_budget_file(source_file, output_file)
