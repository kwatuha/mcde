#!/usr/bin/env python3
"""
Script to process budget Excel file and populate mapping template.
Extracts department names, matches with database, and uses ward to get subcounty.
"""

import pandas as pd
import subprocess
import re
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Optional

# Database connection details
DB_CONTAINER = "gov_db"
DB_NAME = "gov_imbesdb"
DB_USER = "root"

def get_db_password():
    """Get MySQL root password from docker container."""
    try:
        result = subprocess.run(
            ["docker", "exec", DB_CONTAINER, "printenv", "MYSQL_ROOT_PASSWORD"],
            capture_output=True,
            text=True,
            check=True
        )
        return result.stdout.strip() or "root"
    except:
        return "root"

def query_database(query: str) -> List[Dict]:
    """Execute MySQL query and return results as list of dicts."""
    password = get_db_password()
    cmd = [
        "docker", "exec", "-i", DB_CONTAINER,
        "mysql", "-u", DB_USER, f"-p{password}", DB_NAME,
        "-e", query
    ]
    
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            check=True
        )
        # Parse tab-separated output
        lines = result.stdout.strip().split('\n')
        if len(lines) < 2:
            return []
        
        headers = lines[0].split('\t')
        data = []
        for line in lines[1:]:
            if line.strip():
                values = line.split('\t')
                data.append(dict(zip(headers, values)))
        return data
    except subprocess.CalledProcessError as e:
        print(f"Database query error: {e.stderr}")
        return []

def load_database_mappings() -> Tuple[Dict, Dict, Dict]:
    """Load departments, wards, and subcounties from database."""
    print("Loading database mappings...")
    
    # Load departments
    dept_query = "SELECT departmentId, name FROM departments WHERE voided = 0;"
    dept_data = query_database(dept_query)
    departments = {}
    for row in dept_data:
        dept_name = row.get('name', '').strip()
        dept_id = row.get('departmentId', '').strip()
        if dept_name:
            # Store both exact match and normalized versions
            departments[dept_name.lower()] = {'id': dept_id, 'name': dept_name}
            # Also store without "Department:" prefix
            if dept_name.lower().startswith('department:'):
                departments[dept_name.lower().replace('department:', '').strip()] = {'id': dept_id, 'name': dept_name}
    
    # Load wards and subcounties
    ward_query = """
    SELECT w.wardId, w.name as wardName, sc.subcountyId, sc.name as subcountyName 
    FROM wards w 
    LEFT JOIN subcounties sc ON w.subcountyId = sc.subcountyId 
    WHERE w.voided = 0;
    """
    ward_data = query_database(ward_query)
    wards = {}
    subcounties = {}
    
    for row in ward_data:
        ward_name = row.get('wardName', '').strip()
        subcounty_name = row.get('subcountyName', '').strip()
        ward_id = row.get('wardId', '').strip()
        subcounty_id = row.get('subcountyId', '').strip()
        
        if ward_name:
            # Store normalized ward name
            wards[ward_name.lower()] = {
                'id': ward_id,
                'name': ward_name,
                'subcountyId': subcounty_id,
                'subcountyName': subcounty_name
            }
        
        if subcounty_name:
            subcounties[subcounty_name.lower()] = {
                'id': subcounty_id,
                'name': subcounty_name
            }
    
    print(f"Loaded {len(departments)} departments, {len(wards)} wards, {len(subcounties)} subcounties")
    return departments, wards, subcounties

def normalize_text(text: str) -> str:
    """Normalize text for matching."""
    if not text or pd.isna(text):
        return ""
    return str(text).strip().lower()

def find_matching_department(dept_name: str, departments: Dict) -> Optional[Dict]:
    """Find matching department in database."""
    if not dept_name:
        return None
    
    normalized = normalize_text(dept_name)
    
    # Special case: "City" should match "City of Kisumu"
    if normalized == "city":
        for db_dept, dept_info in departments.items():
            if "city" in db_dept and "kisumu" in db_dept:
                return dept_info
    
    # Try exact match
    if normalized in departments:
        return departments[normalized]
    
    # Try partial matches
    for db_dept, dept_info in departments.items():
        if normalized in db_dept or db_dept in normalized:
            return dept_info
    
    # Try removing common prefixes/suffixes
    cleaned = re.sub(r'^(department:|dept\.?|dept\s+)', '', normalized, flags=re.IGNORECASE)
    if cleaned in departments:
        return departments[cleaned]
    
    # Try fuzzy matching for municipalities and special cases
    if "municipality" in normalized or "assembly" in normalized:
        # These might not exist in database, return None
        return None
    
    return None

def find_matching_ward(ward_name: str, wards: Dict) -> Optional[Dict]:
    """Find matching ward in database."""
    if not ward_name:
        return None
    
    # Remove quotes from ward name (e.g., "Nyalenda A" -> Nyalenda A)
    ward_name = re.sub(r'^["\']|["\']$', '', str(ward_name).strip())
    
    normalized = normalize_text(ward_name)
    
    # Handle special cases - "All Wards" or "All Ward" is synonymous to "Countywide"
    # Check for variations: "all wards", "all ward", "all-wards", "all-ward", etc.
    normalized_clean = re.sub(r'[-\s]+', ' ', normalized).strip()
    if normalized_clean in ['all wards', 'all ward']:
        # Return special marker for CountyWide
        return {'name': 'CountyWide', 'subcountyName': 'CountyWide', 'isCountyWide': True}
    
    # Check if it contains "all" and "ward" (in any order)
    words = set(normalized_clean.split())
    if 'all' in words and 'ward' in words:
        # Return special marker for CountyWide
        return {'name': 'CountyWide', 'subcountyName': 'CountyWide', 'isCountyWide': True}
    
    # Handle "countywide" variations
    if 'countywide' in normalized or 'county wide' in normalized_clean:
        # Return special marker for CountyWide
        return {'name': 'CountyWide', 'subcountyName': 'CountyWide', 'isCountyWide': True}
    
    # Handle "City" - might refer to multiple wards in Kisumu Central
    if normalized == 'city':
        return None  # Special case, no specific ward
    
    # Handle compound wards like "Kisumu East and Kisumu Central"
    if ' and ' in normalized:
        return None  # Multiple wards, can't map to single ward
    
    # Try exact match
    if normalized in wards:
        return wards[normalized]
    
    # Try matching with variations (handle "Kabonyo Kanyagwal" vs "KABONYO/KANYAGWAL")
    # Also handle quotes removal (e.g., "Nyalenda A" -> "nyalenda a")
    normalized_clean = normalized.replace('/', ' ').replace('-', ' ').replace("'", '').replace('"', '').replace('  ', ' ').strip()
    for db_ward, ward_info in wards.items():
        db_ward_clean = db_ward.replace('/', ' ').replace('-', ' ').replace("'", '').replace('"', '').replace('  ', ' ').strip()
        if normalized_clean == db_ward_clean:
            return ward_info
    
    # Try matching with quotes removed (e.g., "Nyalenda A" should match "NYALENDA 'A'")
    # Remove quotes and try matching
    normalized_no_quotes = re.sub(r'["\']', '', normalized).strip()
    for db_ward, ward_info in wards.items():
        db_ward_no_quotes = re.sub(r'["\']', '', db_ward).strip()
        if normalized_no_quotes == db_ward_no_quotes:
            return ward_info
    
    # Try partial matches (handle cases like "Central Kisumu" vs "CENTRAL KISUMU")
    for db_ward, ward_info in wards.items():
        if normalized == db_ward:
            return ward_info
        # Try matching key parts (ignore spaces and case)
        if normalized.replace(' ', '').replace('-', '').replace('/', '') == db_ward.replace(' ', '').replace('-', '').replace('/', ''):
            return ward_info
    
    # Try matching individual words (for cases like "Kisumu East" matching "EAST KISUMU" pattern)
    normalized_words = set(normalized.split())
    for db_ward, ward_info in wards.items():
        db_ward_words = set(db_ward.split())
        # If all words match (order independent)
        if normalized_words == db_ward_words:
            return ward_info
    
    return None

def extract_department_from_sheet(df: pd.DataFrame) -> Optional[str]:
    """Extract department name from sheet DataFrame."""
    if df.empty:
        return None
    
    # Check first row for "Department:" pattern
    first_row = df.iloc[0]
    for col in df.columns:
        val = str(first_row[col]) if pd.notna(first_row[col]) else ""
        if 'department:' in val.lower():
            # Extract department name
            match = re.search(r'department:\s*(.+)', val, re.IGNORECASE)
            if match:
                return match.group(1).strip()
    
    return None

def extract_data_from_sheet(df: pd.DataFrame, department: str) -> List[Dict]:
    """Extract project data from a sheet."""
    data = []
    
    # Find header row (usually contains "S/No", "Project", "Ward", "Amount")
    header_row_idx = None
    for idx, row in df.iterrows():
        row_str = ' '.join([str(cell) for cell in row.values if pd.notna(cell)]).lower()
        if 's/no' in row_str or 'project' in row_str or 'ward' in row_str:
            header_row_idx = idx
            break
    
    if header_row_idx is None:
        # Try to find data starting from row 1
        header_row_idx = 0
    
    # Get headers
    headers = df.iloc[header_row_idx]
    
    # Find column indices
    project_col = None
    ward_col = None
    amount_col = None
    
    for idx, header in enumerate(headers):
        header_str = str(header).lower() if pd.notna(header) else ""
        if 'project' in header_str and project_col is None:
            project_col = idx
        elif 'ward' in header_str and ward_col is None:
            ward_col = idx
        elif 'amount' in header_str and amount_col is None:
            amount_col = idx
    
    # If columns not found, try common positions
    if project_col is None:
        project_col = 1 if len(df.columns) > 1 else 0
    if ward_col is None:
        ward_col = 2 if len(df.columns) > 2 else 1
    if amount_col is None:
        amount_col = 3 if len(df.columns) > 3 else 2
    
    # Extract data rows
    for idx in range(header_row_idx + 1, len(df)):
        row = df.iloc[idx]
        
        project = str(row.iloc[project_col]).strip() if pd.notna(row.iloc[project_col]) else ""
        ward = str(row.iloc[ward_col]).strip() if pd.notna(row.iloc[ward_col]) else ""
        amount = row.iloc[amount_col] if pd.notna(row.iloc[amount_col]) else ""
        
        # Skip empty rows
        if not project or project.lower() in ['s/no', 'project', 'ward', 'amount', 'nan']:
            continue
        
        # Skip if amount is not numeric
        try:
            float(amount)
        except (ValueError, TypeError):
            continue
        
        data.append({
            'project': project,
            'ward': ward,
            'amount': amount,
            'department': department
        })
    
    return data

def process_budget_file(source_file: str, template_file: str, output_file: str):
    """Process source budget file and populate template."""
    print(f"Reading source file: {source_file}")
    xls = pd.ExcelFile(source_file)
    
    # Load database mappings
    departments, wards, subcounties = load_database_mappings()
    
    # Process all sheets
    all_data = []
    current_department = None
    
    for sheet_name in xls.sheet_names:
        print(f"\nProcessing sheet: {sheet_name}")
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        
        # Try to extract department from this sheet
        sheet_dept = extract_department_from_sheet(df)
        if sheet_dept:
            current_department = sheet_dept
            print(f"  Found department: {current_department}")
        elif current_department is None:
            print(f"  Warning: No department found and no previous department to continue from")
            continue
        
        # Extract data from sheet
        sheet_data = extract_data_from_sheet(df, current_department)
        print(f"  Extracted {len(sheet_data)} items")
        all_data.extend(sheet_data)
    
    print(f"\nTotal items extracted: {len(all_data)}")
    
    # Create output DataFrame
    output_data = []
    
    for item in all_data:
        # Find matching department
        dept_match = find_matching_department(item['department'], departments)
        db_department = dept_match['name'] if dept_match else "unknown"
        
        # Find matching ward and subcounty
        ward_match = find_matching_ward(item['ward'], wards)
        if ward_match:
            db_ward = ward_match.get('name', 'unknown')
            # For CountyWide, use CountyWide for subcounty too
            if ward_match.get('isCountyWide'):
                db_subcounty = 'CountyWide'
            else:
                db_subcounty = ward_match.get('subcountyName', 'unknown')
        else:
            db_ward = "unknown"
            db_subcounty = "unknown"
        
        output_data.append({
            'BudgetName': 'Approved Budget FY 2025/2026',
            'Department': item['department'],
            'db_department': db_department,
            'Project Name': item['project'],
            'ward': item['ward'],
            'Amount': item['amount'],
            'db_subcounty': db_subcounty,
            'db_ward': db_ward,
            'db_subcounty.1': db_subcounty  # Duplicate column in template
        })
    
    # Create DataFrame
    output_df = pd.DataFrame(output_data)
    
    # Write to template file
    print(f"\nWriting to output file: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='Sheet1', index=False)
        
        # Get the worksheet to adjust column widths and enable wrapping
        worksheet = writer.sheets['Sheet1']
        
        # Define column width settings (column name -> width)
        # Widths are in Excel units (approximately character width)
        column_widths = {
            'BudgetName': 30,
            'Department': 50,  # Long department names
            'db_department': 50,  # Long department names
            'Project Name': 60,  # Long project names
            'ward': 25,
            'Amount': 15,
            'db_subcounty': 30,
            'db_ward': 30,
            'db_subcounty.1': 30
        }
        
        # Calculate optimal widths based on content
        # Get column names from DataFrame
        column_names = list(output_df.columns)
        
        for idx, col_name in enumerate(column_names, start=1):
            col_letter = get_column_letter(idx)
            
            # Calculate max width based on content
            max_length = 0
            if col_name in column_widths:
                # Use predefined width as base
                max_length = column_widths[col_name]
            else:
                # Calculate from content
                max_length = len(str(col_name))  # Header width
                for value in output_df[col_name]:
                    if pd.notna(value):
                        cell_length = len(str(value))
                        if cell_length > max_length:
                            max_length = cell_length
            
            # Set column width (add some padding, max 100 characters)
            adjusted_width = min(max_length + 2, 100)
            worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Enable text wrapping for all cells in this column
            for row in range(2, len(output_df) + 2):  # Start from row 2 (skip header)
                cell = worksheet[f'{col_letter}{row}']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Also wrap header row
        for idx, col_name in enumerate(column_names, start=1):
            col_letter = get_column_letter(idx)
            header_cell = worksheet[f'{col_letter}1']
            header_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
            header_cell.font = Font(bold=True)
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    print(f"Successfully created mapping file with {len(output_df)} rows")
    print(f"\nSummary:")
    print(f"  Departments matched: {len([d for d in output_data if d['db_department'] != 'unknown'])}")
    print(f"  Wards matched: {len([d for d in output_data if d['db_ward'] != 'unknown'])}")
    print(f"  Subcounties matched: {len([d for d in output_data if d['db_subcounty'] != 'unknown'])}")

if __name__ == "__main__":
    source_file = "/home/dev/dev/imes_working/v5/budgets/2025_2026_budgets.xlsx"
    template_file = "/home/dev/dev/imes_working/v5/budgets/budget_mapping_template.xls"
    output_file = "/home/dev/dev/imes_working/v5/budgets/budget_mapping_template.xls"
    
    process_budget_file(source_file, template_file, output_file)
